import io
import os
import re
import time
import threading
import hashlib
import concurrent.futures
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict

import requests
import openpyxl
from PIL import Image
from dateutil import parser as dateutil_parser

IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".webp",
    ".bmp", ".tiff", ".tif", ".heic", ".heif", ".avif",
}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")
SAMPLE_ROWS = 30
MAX_WORKERS = 10
MAX_RETRIES = 4


# ── Column detection ─────────────────────────────────────────────────────────

def _is_image_url(value: str) -> bool:
    if not value.lower().startswith("http"):
        return False
    path = value.split("?")[0].split("#")[0]
    return Path(path).suffix.lower() in IMAGE_EXTENSIONS


def _is_email(value: str) -> bool:
    return bool(EMAIL_RE.match(value))


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    # Skip plain numbers and very short strings — dateutil would misparse them
    if not s or len(s) < 6 or s.replace(".", "", 1).replace("-", "", 1).isdigit():
        return None
    try:
        return dateutil_parser.parse(s, fuzzy=False).replace(tzinfo=None)
    except Exception:
        return None


def _detect_columns(headers: list, sample: list) -> dict:
    """Return best-matching header for 'email', 'image', and 'date' (or None)."""
    n = len(headers)
    email_scores = [0.0] * n
    image_scores = [0.0] * n
    date_scores  = [0.0] * n

    for col in range(n):
        raw    = [row[col] for row in sample if col < len(row)]
        values = [str(v or "").strip() for v in raw]
        non_empty_str = [v for v in values if v]
        non_empty_raw = [r for r in raw if r is not None and str(r).strip()]
        if not non_empty_str:
            continue
        email_scores[col] = sum(_is_email(v) for v in non_empty_str) / len(non_empty_str)
        image_scores[col] = sum(_is_image_url(v) for v in non_empty_str) / len(non_empty_str)
        date_scores[col]  = sum(_parse_date(v) is not None for v in non_empty_raw) / len(non_empty_raw)

    THRESHOLD = 0.5

    def best(scores, exclude=()):
        ranked = sorted(range(n), key=lambda i: -scores[i])
        for idx in ranked:
            if scores[idx] >= THRESHOLD and idx not in exclude:
                return headers[idx]
        return None

    email_col = best(email_scores)
    image_col = best(image_scores)
    email_idx = headers.index(email_col) if email_col else -1
    image_idx = headers.index(image_col) if image_col else -1
    date_col  = best(date_scores, exclude={email_idx, image_idx})

    # Reject a date column where most dates are older than 3 years — likely DOB
    if date_col is not None:
        cutoff     = datetime.now() - timedelta(days=3 * 365)
        dcol_idx   = headers.index(date_col)
        col_dates  = [_parse_date(row[dcol_idx]) for row in sample if dcol_idx < len(row)]
        valid      = [d for d in col_dates if d is not None]
        if not valid or sum(d > cutoff for d in valid) / len(valid) < 0.5:
            date_col = None

    return {"email": email_col, "image": image_col, "date": date_col}


# ── Adaptive download gate ────────────────────────────────────────────────────

class _DownloadGate:
    """Semaphore-based concurrency gate that shrinks on 429 and recovers on success."""

    def __init__(self, initial: int):
        self._sem  = threading.Semaphore(initial)
        self._lock = threading.Lock()
        self._limit       = initial
        self._initial     = initial
        self._held_extra  = 0

    def acquire(self): self._sem.acquire()
    def release(self): self._sem.release()

    def on_rate_limit(self):
        with self._lock:
            if self._limit > 1 and self._sem.acquire(blocking=False):
                self._limit -= 1
                self._held_extra += 1

    def on_success(self):
        with self._lock:
            if self._held_extra > 0:
                self._held_extra -= 1
                self._limit += 1
                self._sem.release()


# ── Per-URL download worker ───────────────────────────────────────────────────

_thread_local = threading.local()

def _get_session() -> requests.Session:
    if not hasattr(_thread_local, "session"):
        s = requests.Session()
        s.headers["User-Agent"] = "Mozilla/5.0"
        _thread_local.session = s
    return _thread_local.session


def _download_one(url: str, gate: _DownloadGate) -> str:
    if not url:
        return "NO_URL"

    session = _get_session()
    for attempt in range(MAX_RETRIES):
        gate.acquire()
        try:
            resp = session.get(url, timeout=20)
        except Exception:
            gate.release()
            time.sleep(min(2 ** attempt, 30))
            continue
        gate.release()

        if resp.status_code == 429:
            gate.on_rate_limit()
            try:
                wait = int(resp.headers["Retry-After"])
            except (KeyError, ValueError, TypeError):
                wait = 2 ** (attempt + 1)
            time.sleep(min(wait, 60))
            continue

        try:
            resp.raise_for_status()
            img = Image.open(io.BytesIO(resp.content)).convert("RGB")
            h = hashlib.sha256(img.tobytes()).hexdigest()
            gate.on_success()
            return h
        except Exception:
            time.sleep(min(2 ** attempt, 30))

    return "DOWNLOAD_FAILED"


# ── GUI ───────────────────────────────────────────────────────────────────────

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Giveaway Deduplicator")
        self.root.resizable(False, False)
        self.root.minsize(460, 0)

        self.filepath    = None
        self.output_path = None
        self._build_ui()

    def _build_ui(self):
        pad = dict(padx=16, pady=6)

        # File picker — ? lives here so it's always visible
        self.file_frame = ttk.LabelFrame(self.root, text="Excel File", padding=8)
        self.file_label = ttk.Label(self.file_frame, text="No file selected", foreground="gray")
        self.file_label.pack(side="left", fill="x", expand=True)
        ttk.Button(self.file_frame, text="?", width=2, command=self._show_help).pack(side="right", padx=(4, 0))
        self.browse_btn = ttk.Button(self.file_frame, text="Browse...", command=self._browse)
        self.browse_btn.pack(side="right")

        # Progress
        self.prog_frame = ttk.Frame(self.root)
        self.progress = ttk.Progressbar(self.prog_frame, mode="determinate")
        self.progress.pack(fill="x")

        self.status_label = ttk.Label(self.root, text="", foreground="gray")

        # Stats area
        self.stats_frame = ttk.LabelFrame(self.root, text="Results", padding=4)
        self.stats_text = tk.Text(
            self.stats_frame,
            height=8,
            state="disabled",
            font=("Courier New", 9),
            relief="flat",
            background="#f5f5f5",
            wrap="none",
        )
        self.stats_text.pack(fill="both", expand=True)

        self.open_btn = ttk.Button(
            self.root, text="Open Output File", command=self._open_output
        )

        # Ordered sections with their pack options — controls visibility
        self._sections = [
            (self.file_frame,   dict(fill="x", **pad)),
            (self.prog_frame,   dict(fill="x", padx=16, pady=(0, 2))),
            (self.status_label, dict(pady=(0, 4))),
            (self.stats_frame,  dict(fill="both", expand=True, padx=16, pady=4)),
            (self.open_btn,     dict(pady=6)),
        ]

        self._set_visible(self.file_frame)

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _ui(self, fn, *args, **kwargs):
        def _safe():
            try:
                fn(*args, **kwargs)
            except tk.TclError:
                pass
        self.root.after(0, _safe)

    def _set_visible(self, *widgets):
        visible = {id(w) for w in widgets}
        for w, _ in self._sections:
            w.pack_forget()
        for w, opts in self._sections:
            if id(w) in visible:
                w.pack(**opts)
        self.root.update_idletasks()
        self.root.geometry("")

    def _set_status(self, text):
        self._ui(self.status_label.config, text=text)

    def _set_progress(self, value, maximum=None):
        kwargs = dict(value=value)
        if maximum is not None:
            kwargs["maximum"] = maximum
        self._ui(self.progress.config, **kwargs)

    def _set_stats(self, text: str):
        def _update():
            self.stats_text.config(state="normal")
            self.stats_text.delete("1.0", "end")
            self.stats_text.insert("1.0", text)
            self.stats_text.config(state="disabled")
        self._ui(_update)

    # ── Browse ────────────────────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        self.filepath = path
        self.file_label.config(text=Path(path).name, foreground="black")
        self._start()

    # ── Help ─────────────────────────────────────────────────────────────────

    def _show_help(self):
        messagebox.showinfo(
            "How it works",
            "This tool removes duplicate giveaway entries based on photo content.\n\n"
            "Rules:\n"
            "• Each unique photo may appear only once in the output.\n"
            "• If the same photo was submitted multiple times (by the same or "
            "different people), only the earliest submission is kept.\n"
            "• One person may have multiple entries as long as each uses a "
            "different photo.\n\n"
            "The tool automatically detects the email, photo URL, and submission "
            "date columns. Photos are compared by their actual pixel content — "
            "the same photo with a different filename or upload date is still "
            "detected as a duplicate. A resized or re-cropped copy is treated "
            "as a different photo.\n\n"
            "If no date column is found, the first row in the file wins ties.\n\n"
            "Your original file is never changed. The result is saved as a new "
            "file with -DEDUPLICATED added to the name."
        )

    # ── Start / run ───────────────────────────────────────────────────────────

    def _start(self):
        self.output_path = None
        self.browse_btn["state"] = "disabled"
        self._set_visible(self.file_frame, self.prog_frame, self.status_label)
        threading.Thread(target=self._deduplicate_safe, daemon=True).start()

    def _deduplicate_safe(self):
        try:
            self._deduplicate()
        except Exception as e:
            self._ui(messagebox.showerror, "Error", str(e))
            self._ui(lambda: self._set_visible(self.file_frame))
        finally:
            self._ui(self.browse_btn.config, state="normal")

    # ── Core logic ────────────────────────────────────────────────────────────

    def _deduplicate(self):
        self._set_status("Reading Excel file...")

        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        rows    = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
        total   = len(rows)

        detected = _detect_columns(headers, rows[:SAMPLE_ROWS])
        if not detected["image"]:
            raise ValueError(
                "Could not find an image URL column. Make sure the file has a "
                "column where values start with 'http' and end with an image extension."
            )

        image_col = headers.index(detected["image"])
        email_col = headers.index(detected["email"]) if detected["email"] else None
        date_col  = headers.index(detected["date"])  if detected["date"]  else None

        # ── Step 1: parallel image download + hash ───────────────────────────
        gate   = _DownloadGate(MAX_WORKERS)
        hashes = [None] * total
        urls   = [str(row[image_col] or "").strip() for row in rows]

        self._set_progress(0, total)
        self._set_status(f"Downloading images (up to {MAX_WORKERS} at a time)...")

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_idx = {
                executor.submit(_download_one, url, gate): i
                for i, url in enumerate(urls)
            }
            done = 0
            for future in concurrent.futures.as_completed(future_to_idx):
                i = future_to_idx[future]
                hashes[i] = future.result()
                done += 1
                self._set_status(
                    f"Downloaded {done} / {total}  (concurrency: {gate._limit})"
                )
                self._set_progress(done)

        # ── Step 2: parse submission dates ───────────────────────────────────
        dates = [None] * total
        if date_col is not None:
            for i, row in enumerate(rows):
                dates[i] = _parse_date(row[date_col])

        # ── Step 3: global dedup by image hash, keep earliest ────────────────
        self._set_status("Deduplicating...")

        hash_to_rows: dict = defaultdict(list)   # hash -> [row indices]
        kept_set:     set  = set()
        dup_count:    dict = {}                   # kept row_idx -> dupes absorbed

        for i in range(total):
            h = hashes[i]
            if h in ("NO_URL", "DOWNLOAD_FAILED"):
                kept_set.add(i)
                dup_count[i] = 0
            else:
                hash_to_rows[h].append(i)

        for h, group in hash_to_rows.items():
            # Sort by date (None last), then by original row order as tiebreak
            group_sorted = sorted(
                group,
                key=lambda i: (dates[i] is None, dates[i] or datetime.min, i),
            )
            keeper = group_sorted[0]
            kept_set.add(keeper)
            dup_count[keeper] = len(group) - 1

        kept_indices = sorted(kept_set)

        # ── Step 4: compute stats ────────────────────────────────────────────
        removed_total = sum(dup_count[i] for i in kept_indices)

        # Per-email: how many of their entries were removed
        email_removed: dict = defaultdict(int)
        if email_col is not None:
            for i in range(total):
                if i not in kept_set:
                    email = str(rows[i][email_col] or "").strip().lower()
                    if email:
                        email_removed[email] += 1

        # Images that appeared from more than one distinct email.
        # Count all removed entries for such groups (including within-account dupes
        # of the same group, since the photo itself was submitted across accounts).
        cross_account_images  = 0
        cross_account_removed = 0
        if email_col is not None:
            for h, group in hash_to_rows.items():
                emails = {str(rows[i][email_col] or "").strip().lower() for i in group}
                if len(emails) > 1:
                    cross_account_images  += 1
                    cross_account_removed += len(group) - 1

        # Most-submitted photo
        most_count        = 0
        most_emails_counts: list = []  # [(email, count), ...]
        if hash_to_rows:
            top_hash  = max(hash_to_rows, key=lambda h: len(hash_to_rows[h]))
            top_group = hash_to_rows[top_hash]
            most_count = len(top_group)
            if email_col is not None:
                tally: dict = defaultdict(int)
                for i in top_group:
                    tally[str(rows[i][email_col] or "").strip().lower()] += 1
                most_emails_counts = sorted(tally.items())

        # ── Step 5: write output ─────────────────────────────────────────────
        self._set_status("Writing output file...")
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Deduplicated"
        out_ws.append(headers + ["duplicates_removed"])

        for i in kept_indices:
            d = dup_count[i]
            out_ws.append(rows[i] + ([d] if d > 0 else [""]))

        p = Path(self.filepath)
        out_path = p.parent / (p.stem + "-DEDUPLICATED" + p.suffix)
        out_wb.save(out_path)
        self.output_path = str(out_path)

        # ── Step 6: build stats text ─────────────────────────────────────────
        kept = len(kept_indices)
        lines = [
            f"{total} entries in  →  {kept} kept,  {removed_total} removed",
            "",
        ]

        if email_removed:
            top = sorted(email_removed.items(), key=lambda x: -x[1])[:3]
            lines.append("Top duplicate submitters:")
            for email, count in top:
                lines.append(f"  {email}  ×{count + 1} submissions of the same photo")
            lines.append("")

        if cross_account_images > 0:
            lines.append(
                f"Same photo across different accounts: "
                f"{cross_account_images} photo(s),  {cross_account_removed} extra entr{'ies' if cross_account_removed != 1 else 'y'} removed"
            )

        if most_count > 1:
            lines.append("")
            lines.append(f"Most submitted photo: {most_count}× total")
            if most_emails_counts:
                parts = ", ".join(f"{email} ({n})" for email, n in most_emails_counts)
                lines.append(f"  Submitted by: {parts}")

        all_dates = [dates[i] for i in range(total) if dates[i] is not None]
        if all_dates:
            first_dt  = min(all_dates)
            last_dt   = max(all_dates)
            span_days = max((last_dt - first_dt).days, 1)
            lines.append("")
            lines.append(
                f"Giveaway period: {first_dt.strftime('%d %b %Y')} – {last_dt.strftime('%d %b %Y')}"
                f"  ({span_days} day{'s' if span_days != 1 else ''})"
            )
            lines.append(
                f"Avg per day:  {total / span_days:.1f} submitted,  {kept / span_days:.1f} unique"
            )

        self._set_stats("\n".join(lines))
        self._ui(lambda: self._set_visible(*[w for w, _ in self._sections]))
        self._set_status("Done.")

    def _open_output(self):
        if self.output_path and Path(self.output_path).exists():
            os.startfile(self.output_path)


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
